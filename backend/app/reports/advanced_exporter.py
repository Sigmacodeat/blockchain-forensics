"""
Advanced Report Exporter
========================

Multi-format export for forensic analysis results

Supported Formats:
- PDF (court-admissible)
- Excel (XLSX)
- CSV (data export)
- JSON (API integration)
- HTML (web viewing)
"""

import logging
import json
import csv
from typing import Dict, List, Optional, Any
from datetime import datetime
from io import BytesIO, StringIO
import hashlib

logger = logging.getLogger(__name__)

# Optional dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available. Install with: pip install openpyxl")


class AdvancedReportExporter:
    """
    Multi-format forensic report exporter
    
    **Features:**
    - Excel export with formatting
    - CSV export for data analysis
    - JSON export for API integration
    - HTML export for web viewing
    - Batch export (all formats)
    
    **Use Cases:**
    - Law enforcement reports
    - Compliance documentation
    - Data analysis (Excel/CSV)
    - API integration (JSON)
    - Web dashboards (HTML)
    """
    
    def __init__(self):
        self.formats_available = {
            'pdf': True,  # Always via pdf_generator.py
            'excel': EXCEL_AVAILABLE,
            'csv': True,  # Built-in
            'json': True,  # Built-in
            'html': True  # Built-in
        }
    
    def export_to_excel(
        self,
        trace_id: str,
        trace_data: Dict,
        findings: Optional[Dict] = None
    ) -> bytes:
        """
        Export trace results to Excel (XLSX)
        
        **Sheets:**
        1. Summary - Overview and key metrics
        2. Transactions - Transaction details table
        3. Addresses - Address list with risk scores
        4. Findings - Risk findings and alerts
        5. Metadata - Trace parameters and timestamps
        
        Args:
            trace_id: Trace ID
            trace_data: Trace analysis results
            findings: Optional findings/alerts
        
        Returns:
            Excel file as bytes
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # 1. Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._build_summary_sheet(ws_summary, trace_id, trace_data, findings)
        
        # 2. Transactions Sheet
        ws_txs = wb.create_sheet("Transactions")
        self._build_transactions_sheet(ws_txs, trace_data)
        
        # 3. Addresses Sheet
        ws_addrs = wb.create_sheet("Addresses")
        self._build_addresses_sheet(ws_addrs, trace_data)
        
        # 4. Findings Sheet
        if findings:
            ws_findings = wb.create_sheet("Findings")
            self._build_findings_sheet(ws_findings, findings)
        
        # 5. Metadata Sheet
        ws_meta = wb.create_sheet("Metadata")
        self._build_metadata_sheet(ws_meta, trace_id, trace_data)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Excel report generated for trace {trace_id}")
        return output.getvalue()
    
    def _build_summary_sheet(self, ws, trace_id: str, trace_data: Dict, findings: Optional[Dict]):
        """Build summary sheet"""
        # Title
        ws['A1'] = "Blockchain Forensic Analysis Report"
        ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
        ws.merge_cells('A1:D1')
        
        # Report Info
        ws['A3'] = "Trace ID:"
        ws['B3'] = trace_id
        ws['A4'] = "Generated:"
        ws['B4'] = datetime.utcnow().isoformat()
        ws['A5'] = "Status:"
        ws['B5'] = trace_data.get('status', 'N/A')
        
        # Bold labels
        for cell in ['A3', 'A4', 'A5']:
            ws[cell].font = Font(bold=True)
        
        # Key Metrics
        ws['A7'] = "KEY METRICS"
        ws['A7'].font = Font(size=14, bold=True, color="1F4E78")
        
        nodes = trace_data.get('graph', {}).get('nodes', {})
        edges = trace_data.get('graph', {}).get('edges', [])
        
        metrics = [
            ("Total Addresses", len(nodes)),
            ("Total Transactions", len(edges)),
            ("High-Risk Addresses", sum(1 for n in nodes.values() if n.get('taint_received', 0) > 0.5)),
            ("Sanctioned Entities", sum(1 for n in nodes.values() if 'OFAC' in n.get('labels', []))),
            ("Total Volume", f"${sum(e.get('amount', 0) for e in edges):,.2f}"),
        ]
        
        row = 8
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Findings Summary
        if findings:
            ws[f'A{row+1}'] = "FINDINGS SUMMARY"
            ws[f'A{row+1}'].font = Font(size=14, bold=True, color="C00000")
            
            row += 2
            for finding in findings.get('alerts', [])[:5]:
                ws[f'A{row}'] = f"⚠️ {finding.get('title', 'Alert')}"
                ws[f'A{row}'].font = Font(color="C00000")
                row += 1
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def _build_transactions_sheet(self, ws, trace_data: Dict):
        """Build transactions sheet"""
        # Headers
        headers = [
            "Transaction Hash",
            "From Address",
            "To Address",
            "Amount",
            "Timestamp",
            "Block Number",
            "Taint Score"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        edges = trace_data.get('graph', {}).get('edges', [])
        for row, edge in enumerate(edges, start=2):
            ws.cell(row, 1, edge.get('tx_hash', 'N/A')[:20] + '...')
            ws.cell(row, 2, edge.get('from', 'N/A')[:15] + '...')
            ws.cell(row, 3, edge.get('to', 'N/A')[:15] + '...')
            ws.cell(row, 4, f"${edge.get('amount', 0):,.2f}")
            ws.cell(row, 5, edge.get('timestamp', 'N/A'))
            ws.cell(row, 6, edge.get('block_number', 'N/A'))
            ws.cell(row, 7, f"{edge.get('taint', 0):.3f}")
        
        # Auto-adjust columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _build_addresses_sheet(self, ws, trace_data: Dict):
        """Build addresses sheet"""
        headers = ["Address", "Risk Score", "Labels", "Taint Received", "Transaction Count", "Total Volume"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        nodes = trace_data.get('graph', {}).get('nodes', {})
        for row, (addr, node) in enumerate(nodes.items(), start=2):
            ws.cell(row, 1, addr[:20] + '...')
            
            risk_score = node.get('taint_received', 0)
            cell = ws.cell(row, 2, f"{risk_score:.3f}")
            
            # Color code by risk
            if risk_score > 0.7:
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif risk_score > 0.4:
                cell.fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
            
            ws.cell(row, 3, ", ".join(node.get('labels', [])))
            ws.cell(row, 4, f"{node.get('taint_received', 0):.3f}")
            ws.cell(row, 5, node.get('tx_count', 0))
            ws.cell(row, 6, f"${node.get('total_volume', 0):,.2f}")
        
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22
    
    def _build_findings_sheet(self, ws, findings: Dict):
        """Build findings sheet"""
        headers = ["Alert Type", "Severity", "Description", "Address", "Details"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        alerts = findings.get('alerts', [])
        for row, alert in enumerate(alerts, start=2):
            ws.cell(row, 1, alert.get('type', 'N/A'))
            
            severity = alert.get('severity', 'medium')
            cell = ws.cell(row, 2, severity.upper())
            
            if severity == 'critical':
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif severity == 'high':
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            
            ws.cell(row, 3, alert.get('description', 'N/A'))
            ws.cell(row, 4, alert.get('address', 'N/A'))
            ws.cell(row, 5, str(alert.get('metadata', {})))
        
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _build_metadata_sheet(self, ws, trace_id: str, trace_data: Dict):
        """Build metadata sheet"""
        ws['A1'] = "Trace Metadata"
        ws['A1'].font = Font(size=14, bold=True)
        
        metadata = [
            ("Trace ID", trace_id),
            ("Chain", trace_data.get('chain', 'N/A')),
            ("Root Address", trace_data.get('root_address', 'N/A')),
            ("Direction", trace_data.get('direction', 'N/A')),
            ("Max Depth", trace_data.get('max_depth', 'N/A')),
            ("Started At", trace_data.get('started_at', 'N/A')),
            ("Completed At", trace_data.get('completed_at', 'N/A')),
            ("Duration (seconds)", trace_data.get('duration_seconds', 'N/A')),
        ]
        
        row = 3
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def export_to_csv(
        self,
        trace_id: str,
        trace_data: Dict,
        entity_type: str = "transactions"
    ) -> str:
        """
        Export to CSV format
        
        Args:
            trace_id: Trace ID
            trace_data: Trace data
            entity_type: 'transactions' or 'addresses'
        
        Returns:
            CSV string
        """
        output = StringIO()
        
        if entity_type == "transactions":
            writer = csv.DictWriter(output, fieldnames=[
                'tx_hash', 'from', 'to', 'amount', 'timestamp', 'block_number', 'taint'
            ])
            writer.writeheader()
            
            for edge in trace_data.get('graph', {}).get('edges', []):
                writer.writerow({
                    'tx_hash': edge.get('tx_hash', ''),
                    'from': edge.get('from', ''),
                    'to': edge.get('to', ''),
                    'amount': edge.get('amount', 0),
                    'timestamp': edge.get('timestamp', ''),
                    'block_number': edge.get('block_number', ''),
                    'taint': edge.get('taint', 0)
                })
        
        elif entity_type == "addresses":
            writer = csv.DictWriter(output, fieldnames=[
                'address', 'risk_score', 'labels', 'taint_received', 'tx_count', 'total_volume'
            ])
            writer.writeheader()
            
            for addr, node in trace_data.get('graph', {}).get('nodes', {}).items():
                writer.writerow({
                    'address': addr,
                    'risk_score': node.get('taint_received', 0),
                    'labels': ','.join(node.get('labels', [])),
                    'taint_received': node.get('taint_received', 0),
                    'tx_count': node.get('tx_count', 0),
                    'total_volume': node.get('total_volume', 0)
                })
        
        csv_content = output.getvalue()
        logger.info(f"CSV export completed for {trace_id} ({entity_type})")
        return csv_content
    
    def export_to_json(
        self,
        trace_id: str,
        trace_data: Dict,
        findings: Optional[Dict] = None,
        pretty: bool = True
    ) -> str:
        """
        Export to JSON format
        
        Args:
            trace_id: Trace ID
            trace_data: Trace data
            findings: Optional findings
            pretty: Pretty-print JSON
        
        Returns:
            JSON string
        """
        export_data = {
            "trace_id": trace_id,
            "generated_at": datetime.utcnow().isoformat(),
            "trace_data": trace_data,
            "findings": findings or {},
            "metadata": {
                "format": "blockchain_forensics_json_v1",
                "generator": "advanced_report_exporter"
            }
        }
        
        indent = 2 if pretty else None
        json_content = json.dumps(export_data, indent=indent, default=str)
        
        logger.info(f"JSON export completed for {trace_id}")
        return json_content
    
    def export_to_html(
        self,
        trace_id: str,
        trace_data: Dict,
        findings: Optional[Dict] = None
    ) -> str:
        """
        Export to HTML format for web viewing
        
        Args:
            trace_id: Trace ID
            trace_data: Trace data
            findings: Optional findings
        
        Returns:
            HTML string
        """
        nodes = trace_data.get('graph', {}).get('nodes', {})
        edges = trace_data.get('graph', {}).get('edges', [])
        
        high_risk_count = sum(1 for n in nodes.values() if n.get('taint_received', 0) > 0.5)
        
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Forensic Report - {trace_id}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 40px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #1F4E78;
            border-bottom: 3px solid #1F4E78;
            padding-bottom: 10px;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 30px 0;
        }}
        .metric {{
            background: #f8f9fa;
            padding: 20px;
            border-left: 4px solid #1F4E78;
        }}
        .metric-label {{
            font-size: 12px;
            color: #666;
            text-transform: uppercase;
        }}
        .metric-value {{
            font-size: 24px;
            font-weight: bold;
            color: #1F4E78;
            margin-top: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th {{
            background: #1F4E78;
            color: white;
            padding: 12px;
            text-align: left;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        .risk-high {{ background-color: #ffe6e6; }}
        .risk-medium {{ background-color: #fff9e6; }}
        .alert {{
            background: #fff3cd;
            border-left: 4px solid #ffc107;
            padding: 15px;
            margin: 10px 0;
        }}
        .alert-critical {{
            background: #f8d7da;
            border-left: 4px solid #dc3545;
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            font-size: 12px;
            color: #666;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🔍 Blockchain Forensic Analysis Report</h1>
        
        <p><strong>Trace ID:</strong> {trace_id}</p>
        <p><strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
        
        <h2>Summary</h2>
        <div class="summary">
            <div class="metric">
                <div class="metric-label">Total Addresses</div>
                <div class="metric-value">{len(nodes)}</div>
            </div>
            <div class="metric">
                <div class="metric-label">Total Transactions</div>
                <div class="metric-value">{len(edges)}</div>
            </div>
            <div class="metric">
                <div class="metric-label">High-Risk Addresses</div>
                <div class="metric-value">{high_risk_count}</div>
            </div>
            <div class="metric">
                <div class="metric-label">Total Volume</div>
                <div class="metric-value">${sum(e.get('amount', 0) for e in edges):,.0f}</div>
            </div>
        </div>
"""
        
        # Findings
        if findings and findings.get('alerts'):
            html += "<h2>⚠️ Findings</h2>\n"
            for alert in findings['alerts'][:10]:
                severity_class = f"alert-{alert.get('severity', 'medium')}"
                html += f"""
        <div class="alert {severity_class}">
            <strong>{alert.get('title', 'Alert')}</strong><br>
            {alert.get('description', 'N/A')}
        </div>
"""
        
        # Top Transactions
        html += """
        <h2>Recent Transactions</h2>
        <table>
            <thead>
                <tr>
                    <th>Transaction Hash</th>
                    <th>From</th>
                    <th>To</th>
                    <th>Amount</th>
                    <th>Taint</th>
                </tr>
            </thead>
            <tbody>
"""
        
        for edge in edges[:20]:
            taint = edge.get('taint', 0)
            row_class = 'risk-high' if taint > 0.7 else ('risk-medium' if taint > 0.4 else '')
            
            html += f"""
                <tr class="{row_class}">
                    <td>{edge.get('tx_hash', 'N/A')[:20]}...</td>
                    <td>{edge.get('from', 'N/A')[:15]}...</td>
                    <td>{edge.get('to', 'N/A')[:15]}...</td>
                    <td>${edge.get('amount', 0):,.2f}</td>
                    <td>{taint:.3f}</td>
                </tr>
"""
        
        html += """
            </tbody>
        </table>
        
        <div class="footer">
            <p><strong>CONFIDENTIAL:</strong> This report contains sensitive forensic analysis data.</p>
            <p>Generated by Blockchain Forensics Platform v1.0</p>
        </div>
    </div>
</body>
</html>
"""
        
        logger.info(f"HTML export completed for {trace_id}")
        return html
    
    def export_all_formats(
        self,
        trace_id: str,
        trace_data: Dict,
        findings: Optional[Dict] = None
    ) -> Dict[str, bytes]:
        """
        Export to all available formats
        
        Returns:
            Dict with format names as keys and file content as values
        """
        exports = {}
        
        # Excel
        if self.formats_available['excel']:
            try:
                exports['excel'] = self.export_to_excel(trace_id, trace_data, findings)
            except Exception as e:
                logger.error(f"Excel export failed: {e}")
        
        # CSV (transactions)
        try:
            exports['csv_transactions'] = self.export_to_csv(
                trace_id, trace_data, "transactions"
            ).encode('utf-8')
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
        
        # CSV (addresses)
        try:
            exports['csv_addresses'] = self.export_to_csv(
                trace_id, trace_data, "addresses"
            ).encode('utf-8')
        except Exception as e:
            logger.error(f"CSV addresses export failed: {e}")
        
        # JSON
        try:
            exports['json'] = self.export_to_json(
                trace_id, trace_data, findings
            ).encode('utf-8')
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
        
        # HTML
        try:
            exports['html'] = self.export_to_html(
                trace_id, trace_data, findings
            ).encode('utf-8')
        except Exception as e:
            logger.error(f"HTML export failed: {e}")
        
        logger.info(f"Batch export completed for {trace_id}: {list(exports.keys())}")
        return exports


# Singleton instance
advanced_exporter = AdvancedReportExporter()
