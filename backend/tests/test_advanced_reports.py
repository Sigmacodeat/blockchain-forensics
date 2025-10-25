"""
Tests for Advanced Reporting System
====================================

Test multi-format report generation
"""

import pytest
from io import BytesIO

from app.reports.advanced_exporter import AdvancedReportExporter, advanced_exporter


@pytest.fixture
def mock_trace_data():
    """Mock trace data for testing"""
    return {
        "trace_id": "test_trace_123",
        "chain": "ethereum",
        "root_address": "0x742d35Cc6634C0532925a3b844Bc454e4438f44e",
        "status": "completed",
        "graph": {
            "nodes": {
                "0x742d35Cc6634C0532925a3b844Bc454e4438f44e": {
                    "address": "0x742d35Cc6634C0532925a3b844Bc454e4438f44e",
                    "taint_received": 1.0,
                    "labels": ["Source"],
                    "tx_count": 5,
                    "total_volume": 10000.0
                },
                "0x1111111111111111111111111111111111111111": {
                    "address": "0x1111111111111111111111111111111111111111",
                    "taint_received": 0.8,
                    "labels": ["High-Risk"],
                    "tx_count": 3,
                    "total_volume": 5000.0
                }
            },
            "edges": [
                {
                    "from": "0x742d35Cc6634C0532925a3b844Bc454e4438f44e",
                    "to": "0x1111111111111111111111111111111111111111",
                    "amount": 5000.0,
                    "tx_hash": "0xabc123...",
                    "timestamp": "2025-01-11T10:00:00Z",
                    "block_number": 12345678,
                    "taint": 0.8
                }
            ]
        }
    }


@pytest.fixture
def mock_findings():
    """Mock findings data"""
    return {
        "alerts": [
            {
                "type": "high_risk_address",
                "severity": "high",
                "title": "High-Risk Address",
                "description": "Test alert",
                "address": "0x1111111111111111111111111111111111111111",
                "metadata": {"risk_score": 0.8}
            }
        ]
    }


def test_csv_export_transactions(mock_trace_data):
    """Test CSV export for transactions"""
    exporter = AdvancedReportExporter()
    
    csv_content = exporter.export_to_csv(
        "test_trace_123",
        mock_trace_data,
        entity_type="transactions"
    )
    
    assert csv_content is not None
    assert "tx_hash" in csv_content
    assert "0xabc123" in csv_content
    assert "5000.0" in csv_content


def test_csv_export_addresses(mock_trace_data):
    """Test CSV export for addresses"""
    exporter = AdvancedReportExporter()
    
    csv_content = exporter.export_to_csv(
        "test_trace_123",
        mock_trace_data,
        entity_type="addresses"
    )
    
    assert csv_content is not None
    assert "address" in csv_content
    assert "risk_score" in csv_content
    assert "0x742d35Cc" in csv_content


def test_json_export(mock_trace_data, mock_findings):
    """Test JSON export"""
    exporter = AdvancedReportExporter()
    
    json_content = exporter.export_to_json(
        "test_trace_123",
        mock_trace_data,
        findings=mock_findings,
        pretty=True
    )
    
    assert json_content is not None
    assert "test_trace_123" in json_content
    assert "trace_data" in json_content
    assert "findings" in json_content
    
    # Verify it's valid JSON
    import json
    data = json.loads(json_content)
    assert data["trace_id"] == "test_trace_123"
    assert len(data["findings"]["alerts"]) == 1


def test_html_export(mock_trace_data, mock_findings):
    """Test HTML export"""
    exporter = AdvancedReportExporter()
    
    html_content = exporter.export_to_html(
        "test_trace_123",
        mock_trace_data,
        findings=mock_findings
    )
    
    assert html_content is not None
    assert "<!DOCTYPE html>" in html_content
    assert "Blockchain Forensic Analysis Report" in html_content
    assert "test_trace_123" in html_content
    assert "High-Risk Address" in html_content


@pytest.mark.skipif(
    not advanced_exporter.formats_available['excel'],
    reason="openpyxl not installed"
)
def test_excel_export(mock_trace_data, mock_findings):
    """Test Excel export"""
    exporter = AdvancedReportExporter()
    
    excel_bytes = exporter.export_to_excel(
        "test_trace_123",
        mock_trace_data,
        findings=mock_findings
    )
    
    assert excel_bytes is not None
    assert len(excel_bytes) > 0
    
    # Verify it's a valid Excel file
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(excel_bytes))
        
        assert "Summary" in wb.sheetnames
        assert "Transactions" in wb.sheetnames
        assert "Addresses" in wb.sheetnames
        assert "Findings" in wb.sheetnames
        assert "Metadata" in wb.sheetnames
        
        # Check Summary sheet has data
        ws = wb["Summary"]
        assert ws["A1"].value == "Blockchain Forensic Analysis Report"
        assert ws["B3"].value == "test_trace_123"
        
    except ImportError:
        pytest.skip("openpyxl not available for validation")


def test_export_all_formats(mock_trace_data, mock_findings):
    """Test batch export to all formats"""
    exporter = AdvancedReportExporter()
    
    exports = exporter.export_all_formats(
        "test_trace_123",
        mock_trace_data,
        findings=mock_findings
    )
    
    assert exports is not None
    assert isinstance(exports, dict)
    
    # Check expected formats
    assert "csv_transactions" in exports
    assert "csv_addresses" in exports
    assert "json" in exports
    assert "html" in exports
    
    # If Excel available, should be included
    if exporter.formats_available['excel']:
        assert "excel" in exports


def test_formats_available():
    """Test format availability check"""
    exporter = AdvancedReportExporter()
    
    assert exporter.formats_available['pdf'] is True
    assert exporter.formats_available['csv'] is True
    assert exporter.formats_available['json'] is True
    assert exporter.formats_available['html'] is True
    # Excel depends on openpyxl installation
    assert isinstance(exporter.formats_available['excel'], bool)


def test_csv_empty_data():
    """Test CSV export with empty data"""
    exporter = AdvancedReportExporter()
    
    empty_trace = {
        "graph": {
            "nodes": {},
            "edges": []
        }
    }
    
    csv_content = exporter.export_to_csv(
        "empty_trace",
        empty_trace,
        entity_type="transactions"
    )
    
    # Should still have header
    assert "tx_hash" in csv_content


def test_json_export_no_findings(mock_trace_data):
    """Test JSON export without findings"""
    exporter = AdvancedReportExporter()
    
    json_content = exporter.export_to_json(
        "test_trace_123",
        mock_trace_data,
        findings=None
    )
    
    import json
    data = json.loads(json_content)
    
    assert data["trace_id"] == "test_trace_123"
    assert data["findings"] == {}


def test_html_export_large_dataset():
    """Test HTML export with large dataset"""
    exporter = AdvancedReportExporter()
    
    # Create large dataset
    large_trace = {
        "graph": {
            "nodes": {f"0x{i:040x}": {
                "address": f"0x{i:040x}",
                "taint_received": 0.5,
                "labels": [],
                "tx_count": 1,
                "total_volume": 100.0
            } for i in range(50)},
            "edges": [{
                "from": f"0x{i:040x}",
                "to": f"0x{i+1:040x}",
                "amount": 100.0,
                "tx_hash": f"0x{i:064x}",
                "timestamp": "2025-01-11T10:00:00Z",
                "block_number": 12345678 + i,
                "taint": 0.5
            } for i in range(49)]
        }
    }
    
    html_content = exporter.export_to_html(
        "large_trace",
        large_trace
    )
    
    # Should limit to 20 transactions in display
    assert html_content is not None
    assert len(html_content) > 0


@pytest.mark.asyncio
async def test_report_api_formats_endpoint(client):
    """Test GET /api/v1/reports/formats endpoint"""
    response = client.get("/api/v1/reports/formats")
    
    assert response.status_code == 200
    data = response.json()
    
    assert "formats" in data
    assert "pdf" in data["formats"]
    assert "excel" in data["formats"]
    assert "recommendations" in data
    assert data["recommendations"]["legal"] == "pdf"


@pytest.mark.asyncio
async def test_report_api_metadata_endpoint(client):
    """Test GET /api/v1/reports/metadata/{trace_id} endpoint"""
    response = client.get("/api/v1/reports/metadata/test_trace_123")
    
    assert response.status_code == 200
    data = response.json()
    
    assert "trace_id" in data
    assert "statistics" in data
    assert "available_formats" in data
